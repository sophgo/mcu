#!/usr/bin/python3

from openpyxl import load_workbook
import sys, os
from argparse import ArgumentParser

def error(msg):
    sys.stderr.write(msg)

def indent(filename):
    cmd = 'indent -nbad -bap -nbc -bbo -hnl -br -brs -c33 -cd33 -ncdb -ce -ci4 -cli0 -d0 -di1 -nfc1 -i8 -ip0 -l80 -lp -npcs -nprs -npsl -sai -saf -saw -ncs -nsc -sob -nfca -cp33 -ss -ts8 -il1 {}'.format(filename)
    os.system(cmd)
    # remove temperary file
    os.system('rm -rf {}~'.format(filename))

class Pin:
    def __init__(self, data):
        self.data = data
        port = None
        pin = None
        pin_name = data['pin name']
        if pin_name[0] == 'P' and ord('A') <= ord(pin_name[1]) <= ord('H'):
            port = pin_name[1]
            pin = pin_name[2]
            if len(pin_name) > 3 and ord('0') <= ord(pin_name[3]) <= ord('9'):
                pin = pin + pin_name[3]
            pin = int(pin)
        self.port = port
        self.pin = pin
        self.net_name = data['net name']

    def __str__(self):
        if (self.get_port()):
            return self.get_port() + str(self.get_pin()) + ': ' + self.get_net_name()
        else:
            return self.get_net_name()

    def get_port(self):
        return self.port
    def get_pin(self):
        return self.pin

    def get_pin_name(self):
        return '{}_PIN'.format(self.get_code_name())

    def get_net_name(self):
        return self.net_name
    def get_code_name(self):
        return self.get_net_name().replace('/', '_').replace('.', '_')

    def get_od(self):
        if self.data['default'] == 0:
            return 0
        elif self.data['default'] == 1:
            return 1
        else:
            return None

    def get_otype(self):
        # 0 push-pull
        # 1 open-drain
        if self.data['output type'] == 'OD':
            return 'GPIO_OTYPE_OD'
        elif self.data['output type'] == 'PP':
            return 'GPIO_OTYPE_PP'
        else:
            return None

    def get_mode(self):
        # 0b00 input
        # 0b01 output
        # 0b10 alternate function
        # 0b11 analog
        if self.data['mode'] == 'INPUT':
            return 'GPIO_MODE_INPUT'
        elif self.data['mode'] == 'OUTPUT':
            return 'GPIO_MODE_OUTPUT'
        elif self.data['mode'] == 'AF':
            return 'GPIO_MODE_AF'
        elif self.data['mode'] == 'ANALOG':
            return 'GPIO_MODE_ANALOG'
        else:
            return None
    
    def get_ospeed(self):
        # 0b00 low
        # 0b01 medium
        # 0b10 high
        # 0b11 very high
        if self.data['output speed'] == 'LOW':
            return 'GPIO_OSPEED_LOW'
        elif self.data['output speed'] == 'MEDIUM':
            return 'GPIO_OSPEED_MED'
        elif self.data['output speed'] == 'HIGH':
            return 'GPIO_OSPEED_HIGH'
        elif self.data['output speed'] == 'VERY-HIGH':
            return 'GPIO_OSPEED_VERYHIGH'
        else:
            return None

    def get_pupd(self):
        # 0b00 no pull-up, pull-down
        # 0b01 pull-up
        # 0b10 pull-down
        # 0b11 reserved
        if self.data['pull'] == 'NO-PULL':
            return 'GPIO_PUPD_NONE'
        elif self.data['pull'] == 'PULL-UP':
            return 'GPIO_PUPD_PULLUP'
        elif self.data['pull'] == 'PULL-DOWN':
            return 'GPIO_PUPD_PULLDOWN'
        else:
            return None

    def get_af(self):
        if type(self.data['af']) is int:
            return "GPIO_AF{}".format(self.data['af'])
        elif type(self.data['af']) is str:
            try:
                return "GPIO_AF{}".format(int(self.data['af']))
            except:
                error('{} Invalid AF, AF should be a number\n'.format(self.data['pin name']))
                raise Exception()
                return None
        else:
            return None

class Port:
    def __init__(self, name):
        self.name = name
        self.pins = []
    def add_pin(self, pin):
        num = pin.get_pin()
        if (num != None):
            if (len(self.pins) > num):
                self.pins[pin.get_pin()] = pin
            else:
                for i in range(num - len(self.pins)):
                    self.pins.append(None)
                self.pins.append(pin)
        else:
            self.pins.append(pin)
    def __str__(self):
        s = self.name + ': '
        for p in self.pins:
            if p:
                pin_num = p.get_pin()
                if pin_num != None:
                    pin_num = str(pin_num) + ':'
                else:
                    pin_num = ''
                s += '(' + pin_num + p.get_net_name() + ') '
            else:
                s += '(' + 'NONE' + ')'
        s += '\n'
        return s

    def __bool__(self):
        for pin in self.pins:
            if pin:
                return True
        return False

    def gen_code(self):

        code = '\t/* port {} configuration */\n'.format(self.name)

        # setup output default value
        code += '\t/* output data, default output state */\n'
        reg = ''
        for pin in self.pins:
            if not pin:
                continue
            val = pin.get_od()
            if val == None or val == 0:
                continue
            reg += ' {} |'.format(pin.get_pin_name())
        if reg == '':
            reg = ' 0'
        else:
            reg = reg.rstrip(' |')

        code += 'GPIO_ODR(GPIO{}) ={};\n'.format(self.name, reg)

        # push pull or open drain
        code += '\t/* output type, push-pull or open-drain */\n'
        reg = ''
        for pin in self.pins:
            if not pin:
                continue
            val = pin.get_otype()
            if val == None or val == 'GPIO_OTYPE_PP':
                continue
            reg += ' ({} << {}) |'.format(val, pin.get_pin())
        if reg == '':
            reg = ' 0'
        else:
            reg = reg.rstrip(' |')

        code += 'GPIO_OTYPER(GPIO{}) ={};\n'.format(self.name, reg)

        # output speed
        code += '\t/* output speed, low, medium, high, very high */\n'
        reg = ''
        mask = 0

        for pin in self.pins:
            if not pin:
                continue
            if pin.get_ospeed() != None:
                mask |= 3 << (pin.get_pin() * 2)

        for pin in self.pins:
            if not pin:
                continue
            val = pin.get_ospeed()
            if val == None:
                continue
            reg += ' ({} << ({} * 2)) |'.format(val, pin.get_pin())
        if reg == '':
            reg = ' 0'
        else:
            reg = reg.rstrip(' |')

        # skip configuration, if no fields should modify
        if not mask == 0:
            code += 'GPIO_OSPEEDR(GPIO{}) = (GPIO_OSPEEDR(GPIO{}) & ~0x{:08x}) |{};\n'.format(self.name, self.name, mask, reg)
        else:
            code += '\t/* reset state */\n'

        # pull-up pull-down
        code += '\t/* pull-up pull-down */\n'
        reg = ''
        mask = 0

        for pin in self.pins:
            if not pin:
                continue
            if pin.get_pupd() != None:
                mask |= 3 << (pin.get_pin() * 2)

        for pin in self.pins:
            if not pin:
                continue
            val = pin.get_pupd()
            if val == None:
                continue
            reg += ' ({} << ({} * 2)) |'.format(val, pin.get_pin())
        if reg == '':
            reg = ' 0'
        else:
            reg = reg.rstrip(' |')

        # skip configuration, if no fields should modify
        if not mask == 0:
            code += 'GPIO_PUPDR(GPIO{}) = (GPIO_PUPDR(GPIO{}) & ~0x{:08x}) |{};\n'.format(self.name, self.name, mask, reg)
        else:
            code += '\t/* reset state */\n'

        # af
        code += '\t/* af, alternative function selection */\n'
        reg_low = ''
        reg_high = ''
        mask_low = 0
        mask_high = 0

        for pin in self.pins:
            if not pin:
                continue
            # if pin.get_af() != None:
            #     print('GPIO{}{} AF{}'.format(pin.get_port(), pin.get_pin(), pin.get_af()))
            if pin.get_af() != None:
                if pin.get_pin() < 8:
                    mask_low |= 0xf << (pin.get_pin() * 4)
                else:
                    mask_high |= 0xf << ((pin.get_pin() - 8) * 4)

        for pin in self.pins:
            if not pin:
                continue
            val = pin.get_af()
            if val == None:
                continue
            if pin.get_pin() < 8:
                reg_low += ' ({} << ({} * 4)) |'.format(val, pin.get_pin())
            else:
                reg_high += ' ({} << (({} - 8) * 4)) |'.format(val, pin.get_pin())

        if reg_low == '':
            reg_low = ' 0'
        else:
            reg_low = reg_low.rstrip(' |')

        if reg_high == '':
            reg_high = ' 0'
        else:
            reg_high = reg_high.rstrip(' |')

        # skip configuration, if no fields should modify
        if not mask_low == 0:
            code += 'GPIO_AFRL(GPIO{}) = (GPIO_AFRL(GPIO{}) & ~0x{:08x}) | {};\n'.format(self.name, self.name, mask_low, reg_low)
        else:
            code += '\t/* AFRL reset state */\n'

        if not mask_high == 0:
            code += 'GPIO_AFRH(GPIO{}) = (GPIO_AFRH(GPIO{}) & ~0x{:08x}) | {};\n'.format(self.name, self.name, mask_high, reg_high)
        else:
            code += '\t/* AFRH reset state */\n'

        # mode
        code += '\t/* mode, input, output, alternate function or analog */\n'
        reg = ''
        mask = 0

        for pin in self.pins:
            if not pin:
                continue
            if pin.get_mode() != None:
                mask |= 3 << (pin.get_pin() * 2)

        for pin in self.pins:
            if not pin:
                continue
            val = pin.get_mode()
            if val == None:
                continue
            reg += ' ({} << ({} * 2)) |'.format(val, pin.get_pin())
        if reg == '':
            reg = ' 0'
        else:
            reg = reg.rstrip(' |')

        # skip configuration, if no fields should modify
        if not mask == 0:
            code += 'GPIO_MODER(GPIO{}) = (GPIO_MODER(GPIO{}) & ~0x{:08x}) | {};\n'.format(self.name, self.name, mask, reg)
        else:
            code += '\t/* reset state */\n'

        return code

    def gen_header(self):

        code = '/* port {} definition */\n'.format(self.name)
        for pin in self.pins:
            if not pin:
                continue
            code += '#define {}_PORT\t\tGPIO{}\n'.format(pin.get_code_name(), self.name)
            code += '#define {}_PIN\t\tGPIO_PIN_{}\n'.format(pin.get_code_name(), pin.get_pin())
            code += '#define {}_EXTI\t\tEXTI_{}\n'.format(pin.get_code_name(), pin.get_pin())
        return code

class PinList:
    def __init__(self, filename = 'pin.xlsx', sheetname = 'pin'):
        self.book_name = filename
        self.sheet_name = sheetname

        self.book = load_workbook(filename = filename)
        self.sheet = self.book[sheetname]

        self.ports = []
        port = 'A'
        while (ord(port) <= ord('H')):
            self.ports.append(Port(port))
            port = chr(ord(port) + 1)

    def load(self):
        sheet = self.sheet
        self.cols = {
            'pin number': 'Z',
            'pin name': 'Z',
            'net name': 'Z',
            'mode': 'Z',
            'af': 'Z',
            'pull': 'Z',
            'output type': 'Z',
            'output speed': 'Z',
            'default': 'Z',
            'description': 'Z',
        }
        # search the first row -- table header
        for col in range(ord('A'), ord('Z')):
            cell = '{}{}'.format(chr(col), 1)
            value = sheet[cell].value

            if str(value).lower() in self.cols:
                self.cols[str(value).lower()] = chr(col)
        for key, value in self.cols.items():
            if value == 'Z':
                error(key, 'not found')
                return False
        # phase pins
        row = 2
        while (True):
            pin_number_cell = '{}{}'.format(self.cols['pin number'], row)
            pin_number_value = sheet[pin_number_cell].value
            if (not pin_number_value):
                break;

            pin_data = {}
            pin_data['pin number'] = pin_number_value

            cell = '{}{}'.format(self.cols['pin name'], row)
            pin_data['pin name'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['net name'], row)
            pin_data['net name'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['mode'], row)
            pin_data['mode'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['af'], row)
            pin_data['af'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['pull'], row)
            pin_data['pull'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['output type'], row)
            pin_data['output type'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['output speed'], row)
            pin_data['output speed'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['default'], row)
            pin_data['default'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['description'], row)
            pin_data['description'] = sheet[cell].value

            pin = Pin(pin_data)
    
            if pin_data['net name'] and pin_data['net name'].lower() != 'nc':
                port = pin.get_port()
                num = pin.get_pin()
                if port:
                    self.ports[ord(port) - ord('A')].add_pin(pin)

            row = row + 1

        return True;
    def __str__(self):
        s = ''
        for port in self.ports:
            s = s + str(port) + '\n'
        return s
    def gen_code(self):
        code = '/* THIS IS AUTO GENERATED CODE */\n\n'
        code += '#include <pin.h>\n\n'
        code += 'void pin_init(void)\n'
        code += '{\n'

        for port in self.ports:
            if port:
                code += port.gen_code() + '\n'
        # code += self.ports[0].gen_code() + '\n'

        code += '}\n'
        code += '/* AUTO GENERATED CODE END */'
        return code
    def gen_header(self):
        code = '/* THIS IS AUTO GENERATED CODE */\n\n'
        code += '#ifndef __PIN_H__\n'
        code += '#define __PIN_H__\n\n'
        code += '#include <gd32e50x_gpio.h>\n\n'
        for port in self.ports:
            if port:
                code += port.gen_header() + '\n'

        code += '\n#endif\n'
        code += '/* AUTO GENERATED CODE END */'
        return code

class Node:
    def __init__(self, name, node_type, delay, net_name):
        self.name = name
        self.type = node_type
        self.delay = delay
        self.net_name = net_name
    def __str__(self):
        return '{} {} {} {}us'.format(self.name, self.net_name, self.type, self.delay)
    def get_code_name(self):
        if not self.net_name:
            return None

        tmp = self.net_name.replace('/', '_').replace('.', '_').replace(' ', '_')
        if self.type == 'FUNCTION':
            tmp = tmp.lower()
        return tmp 

    def gen_code(self):
        code = '{{\"{}\", NODE_TYPE_{}, {},\n'.format(self.name, self.type, self.delay)

        code_name = self.get_code_name();
        if self.type == 'ENABLE' or self.type == 'CHECK':
            p0 = code_name + '_PORT'
            p1 = code_name + '_PIN'
        else:
            if code_name:
                p0 = code_name + '_on'
                p1 = code_name + '_off'
            else:
                p0 = 'NULL'
                p1 = 'NULL'
        code += '{{(unsigned long){}, (unsigned long){}}},\n'.format(p0, p1)

        code += '},\n'
        return code

    def gen_porting_func_dec(self):
        if self.type == 'ENABLE' or self.type == 'CHECK':
            return ''
        s = ''
        code_name = self.get_code_name()
        if not code_name:
            return ''
        p0 = code_name + '_on'
        p1 = code_name + '_off'
        s += 'int {}(void);\n'.format(p0)
        s += 'void {}(void);\n'.format(p1)
        return s

    def gen_porting_func_imp(self):
        if self.type == 'ENABLE' or self.type == 'CHECK':
            return ''
        s = ''
        code_name = self.get_code_name()
        if not code_name:
            return ''
        p0 = code_name + '_on'
        p1 = code_name + '_off'
        s += 'int __weak {}(void)\n'.format(p0)
        s += '{\n'
        s += '\t/* add customer code here */\n'
        s += '\treturn 0;\n'
        s += '}\n'
        s += 'void __weak {}(void)\n'.format(p1)
        s += '{\n'
        s += '\t/* add customer code here */\n'
        s += '}\n'
        return s

class Power:
    def __init__(self, filename = 'pin.xlsx', sheetname = 'pin'):
        self.book_name = filename
        self.sheet_name = sheetname

        self.book = load_workbook(filename = filename)
        self.sheet = self.book[sheetname]
        self.nodes = []

    def load(self):
        sheet = self.sheet
        self.cols = {
            'name': 'Z',
            'net name': 'Z',
            'type': 'Z',
            'delay': 'Z',
        }
        # search the first row -- table header
        for col in range(ord('A'), ord('Z')):
            cell = '{}{}'.format(chr(col), 1)
            value = sheet[cell].value

            if str(value).lower() in self.cols:
                self.cols[str(value).lower()] = chr(col)
        for key, value in self.cols.items():
            if value == 'Z':
                error('{} not found'.format(key))
                return False
        # phase pins
        row = 2
        while (True):
            name_cell = '{}{}'.format(self.cols['name'], row)
            name_value = sheet[name_cell].value
            if (not name_value):
                break;

            node_data = {}
            node_data['name'] = name_value

            cell = '{}{}'.format(self.cols['net name'], row)
            node_data['net name'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['type'], row)
            node_data['type'] = sheet[cell].value

            cell = '{}{}'.format(self.cols['delay'], row)
            node_data['delay'] = sheet[cell].value

            self.nodes.append(Node( node_data['name'], \
                                    node_data['type'], \
                                    node_data['delay'], \
                                    node_data['net name']))

            row = row + 1

        return True;
    def __str__(self):
        s = ''
        for node in self.nodes:
            s = s + str(node) + '\n'
        return s

    def gen_code(self):
        code = '/* AUTO GENERATED CODE */\n\n'

        code += '#include <power.h>\n'
        code += '#include <pin.h>\n'
        code += '#include <common.h>\n'
        code += '#include <stdlib.h>\n\n'

        for node in self.nodes:
            code += node.gen_porting_func_dec() + '\n'

        code += 'struct power_node const board_power_nodes[{}] = {{\n\n'.format(len(self.nodes))

        for node in self.nodes:
            code += node.gen_code() + '\n'

        code += '\n};\n'

        code += '\n/* AUTO GENERATED CODE END */'
        return code

    def gen_porting(self):
        code = '/* AUTO GENERATED CODE */\n\n'
        code += '#ifndef __weak\n'
        code += '#define __weak __attribute__((weak))\n'
        code += '#endif\n\n'
        for node in self.nodes:
            code += node.gen_porting_func_imp()
        code += '\n/* AUTO GENERATED CODE END */'
        return code

    def gen_header(self):
        code = '/* AUTO GENERATED CODE */\n\n'
        code = '#ifndef __BOARD_POWER_H__\n'
        code += '#define __BOARD_POWER_H__\n\n'
        code += 'extern struct power_node const board_power_nodes[{}];\n\n'.format(len(self.nodes))
        code += '#endif\n'
        code += '\n/* AUTO GENERATED CODE END */'
        return code


def parse_pin_config(args):
    config_type = args['type']
    source_file = args['source']
    header_file = args['header']
    config_file = args['configuration-file']
    porting_file = args['porting']
    sheet = args['sheet']

    if not sheet:
        sheet = 'pin'

    pin_list = PinList(config_file, sheet)
    if not pin_list.load():
        # print('book:', pin_list.book_name, ', sheet:', pin_list.sheet_name, ', load failed')
        error('loading configuration file failed')
        return 1
    else:
        # print('book:', pin_list.book_name, ', sheet:', pin_list.sheet_name, ', load success')
        pass

    if source_file:
        source = pin_list.gen_code()
        f = open(source_file, 'w')
        if not f:
            error('cannot open source file for writing')
            return 2
        f.write(source)
        f.close()
        indent(source_file)

    if header_file:
        header = pin_list.gen_header()
        f = open(header_file, 'w')
        if not f:
            error('cannot open header file for writing')
            return 3
        f.write(header)
        f.close()
        indent(header_file)
    return 0

def parse_power_config(args):
    config_type = args['type']
    source_file = args['source']
    header_file = args['header']
    config_file = args['configuration-file']
    porting_file = args['porting']
    sheet = args['sheet']

    if not sheet:
        sheet = 'power'

    power = Power(config_file, sheet)
    if not power.load():
        return 1

    if source_file:
        source = power.gen_code()
        f = open(source_file, 'w')
        if not f:
            error('cannot open source file for writing')
            return 2
        f.write(source)
        f.close()
        indent(source_file)

    if header_file:
        header = power.gen_header()
        f = open(header_file, 'w')
        if not f:
            error('cannot open header file for writing')
            return 3
        f.write(header)
        f.close()
        indent(header_file)

    if porting_file:
        porting = power.gen_porting()
        f = open(porting_file, 'w')
        if not f:
            error('cannot open porting file for writing')
            return 2
        f.write(porting)
        f.close()
        indent(porting_file)

    return 0


if __name__ == '__main__':

    ######################################

    parser = ArgumentParser(description = 'STM32 GPIO configuration generate tool, generate c language source code and header file')
    parser.add_argument('-y, --type', help = 'pin or power, default is pin', dest = 'type', type = str, action = 'store', default = 'pin')
    parser.add_argument('-s, --source', help = 'output source code to specified c source file', dest = 'source', type = str, action = 'store', default = None)
    parser.add_argument('-e, --header', help = 'output pin definition to specified c header file', dest = 'header', type = str, action = 'store', default = None)
    parser.add_argument('-t, --sheet', help = 'sheet, default is \'pin\' if pin config, \'power if power config\'', dest = 'sheet', type = str, action = 'store', default = None)
    parser.add_argument('-p, --porting', help = 'output source code that customer should implement', dest = 'porting', type = str, action = 'store', default = None)
    parser.add_argument('configuration-file', help = 'pin configuration file with microsoft excel format(\'xlsx\' as extension)', type = str, action = 'store')

    args = vars(parser.parse_args())

    if args['type'] == 'pin':
        err = parse_pin_config(args)
    else:
        err = parse_power_config(args)

    sys.exit(err)

